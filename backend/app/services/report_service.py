"""
Report generation service.
PDF: WeasyPrint (HTML → PDF, professional layout)
XLSX: openpyxl (formatted spreadsheet with column widths, headers)
"""
import io
from calendar import monthrange
from datetime import UTC, datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.attendance import AttendanceRecord, Session
from app.models.inventory import (
    BorrowTransaction,
    BorrowTransactionItem,
    Equipment,
    EquipmentCondition,
    TransactionStatus,
)
from app.models.user import User


HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")   # OSCA navy blue
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill("solid", fgColor="EBF0F7")


class ReportService:
    def __init__(self, db: AsyncSession) -> None:
        self.db = db

    # ── Attendance PDF ─────────────────────────────────────────────────────────

    async def generate_attendance_pdf(
        self,
        sport_or_art: str | None = None,
        date_from: datetime | None = None,
        date_to: datetime | None = None,
    ) -> bytes:
        records = await self._fetch_attendance_records(sport_or_art, date_from, date_to)
        html = self._render_attendance_html(records, sport_or_art, date_from, date_to)
        return self._html_to_pdf(html)

    async def generate_attendance_xlsx(
        self,
        sport_or_art: str | None = None,
        date_from: datetime | None = None,
        date_to: datetime | None = None,
    ) -> bytes:
        records = await self._fetch_attendance_records(sport_or_art, date_from, date_to)
        return self._build_attendance_xlsx(records)

    # ── Inventory PDF ──────────────────────────────────────────────────────────

    async def generate_inventory_pdf(self) -> bytes:
        equipment = await self._fetch_all_equipment()
        html = self._render_inventory_html(equipment)
        return self._html_to_pdf(html)

    async def generate_inventory_xlsx(self) -> bytes:
        equipment = await self._fetch_all_equipment()
        return self._build_inventory_xlsx(equipment)

    # ── Inventory Monthly Summary ──────────────────────────────────────────────

    async def generate_inventory_monthly_report(self, year: int, month: int) -> dict:
        """
        Build a monthly inventory summary dict.
        Covers borrows started OR returned in the given month.
        """
        from datetime import timezone
        # Month boundaries (UTC)
        _, last_day = monthrange(year, month)
        period_start = datetime(year, month, 1, tzinfo=timezone.utc)
        period_end = datetime(year, month, last_day, 23, 59, 59, tzinfo=timezone.utc)

        # Total active equipment
        total_active = (await self.db.execute(
            select(func.count(Equipment.id)).where(Equipment.is_active == True)
        )).scalar_one()

        # Borrowed this month (transactions whose borrowed_at falls in period)
        borrowed_this_month = (await self.db.execute(
            select(func.count(BorrowTransaction.id)).where(
                BorrowTransaction.borrowed_at >= period_start,
                BorrowTransaction.borrowed_at <= period_end,
            )
        )).scalar_one()

        # Returned this month
        returned_this_month = (await self.db.execute(
            select(func.count(BorrowTransaction.id)).where(
                BorrowTransaction.returned_at >= period_start,
                BorrowTransaction.returned_at <= period_end,
            )
        )).scalar_one()

        # Still overdue at end of month
        overdue_at_end = (await self.db.execute(
            select(func.count(BorrowTransaction.id)).where(
                BorrowTransaction.status == TransactionStatus.OVERDUE,
            )
        )).scalar_one()

        # Top 5 most-borrowed equipment in period
        top_eq_result = await self.db.execute(
            select(Equipment.name, func.count(BorrowTransactionItem.id).label("borrow_count"))
            .join(BorrowTransactionItem, BorrowTransactionItem.equipment_id == Equipment.id)
            .join(BorrowTransaction, BorrowTransactionItem.transaction_id == BorrowTransaction.id)
            .where(
                BorrowTransaction.borrowed_at >= period_start,
                BorrowTransaction.borrowed_at <= period_end,
            )
            .group_by(Equipment.name)
            .order_by(func.count(BorrowTransactionItem.id).desc())
            .limit(5)
        )
        top_equipment = [
            {"name": row.name, "borrow_count": row.borrow_count}
            for row in top_eq_result.all()
        ]

        # Condition breakdown
        condition_result = await self.db.execute(
            select(Equipment.condition, func.count(Equipment.id).label("count"))
            .where(Equipment.is_active == True)
            .group_by(Equipment.condition)
        )
        condition_breakdown = {
            row.condition.value: row.count for row in condition_result.all()
        }

        return {
            "period": {"year": year, "month": month},
            "total_active_equipment": total_active,
            "borrowed_this_month": borrowed_this_month,
            "returned_this_month": returned_this_month,
            "overdue_at_end_of_month": overdue_at_end,
            "top_5_borrowed": top_equipment,
            "condition_breakdown": condition_breakdown,
            "generated_at": datetime.now(UTC).isoformat(),
        }

    async def render_monthly_report_pdf(self, report: dict) -> bytes:
        html = self._render_monthly_html(report)
        return self._html_to_pdf(html)

    async def render_monthly_report_xlsx(self, report: dict) -> bytes:
        return self._build_monthly_xlsx(report)

    # ── Dashboard Summary ──────────────────────────────────────────────────────

    async def get_dashboard_summary(self) -> dict:
        # Total students
        total_students = (await self.db.execute(
            select(func.count(User.id)).where(User.role == "student", User.is_active == True)
        )).scalar_one()

        # Enrolled with face
        face_enrolled = (await self.db.execute(
            select(func.count(User.id)).where(
                User.role == "student", User.is_active == True, User.is_face_enrolled == True
            )
        )).scalar_one()

        # Today's attendance
        today = datetime.now(UTC).date()
        today_attendance = (await self.db.execute(
            select(func.count(AttendanceRecord.id)).where(
                func.date(AttendanceRecord.time_in) == today
            )
        )).scalar_one()

        # Equipment stats
        total_equipment = (await self.db.execute(
            select(func.sum(Equipment.total_quantity)).where(Equipment.is_active == True)
        )).scalar_one() or 0

        borrowed_equipment = (await self.db.execute(
            select(func.sum(Equipment.total_quantity - Equipment.available_quantity)).where(
                Equipment.is_active == True
            )
        )).scalar_one() or 0

        # Overdue transactions
        overdue_count = (await self.db.execute(
            select(func.count(BorrowTransaction.id)).where(
                BorrowTransaction.status == TransactionStatus.OVERDUE
            )
        )).scalar_one()

        return {
            "students": {
                "total": total_students,
                "face_enrolled": face_enrolled,
                "enrollment_rate": round(face_enrolled / total_students * 100, 1) if total_students else 0,
            },
            "attendance": {
                "today": today_attendance,
            },
            "equipment": {
                "total": int(total_equipment),
                "borrowed": int(borrowed_equipment),
                "available": int(total_equipment - borrowed_equipment),
            },
            "transactions": {
                "overdue": overdue_count,
            },
            "generated_at": datetime.now(UTC).isoformat(),
        }

    # ── Private Helpers ────────────────────────────────────────────────────────

    async def _fetch_attendance_records(
        self,
        sport_or_art: str | None,
        date_from: datetime | None,
        date_to: datetime | None,
    ) -> list[dict]:
        query = (
            select(AttendanceRecord, User, Session)
            .join(User, AttendanceRecord.student_id == User.id)
            .join(Session, AttendanceRecord.session_id == Session.id)
        )
        if sport_or_art:
            query = query.where(Session.sport_or_art == sport_or_art)
        if date_from:
            query = query.where(AttendanceRecord.time_in >= date_from)
        if date_to:
            query = query.where(AttendanceRecord.time_in <= date_to)
        query = query.order_by(AttendanceRecord.time_in.desc())

        result = await self.db.execute(query)
        rows = []
        for record, user, session in result.all():
            rows.append({
                "student_name": user.full_name,
                "student_id": user.student_id,
                "session_name": session.name,
                "sport_or_art": session.sport_or_art,
                "activity_type": session.activity_type.value,
                "time_in": record.time_in,
                "time_out": record.time_out,
                "duration_minutes": record.duration_minutes,
                "confidence": record.time_in_confidence,
                "is_complete": record.is_complete,
            })
        return rows

    async def _fetch_all_equipment(self) -> list[dict]:
        result = await self.db.execute(
            select(Equipment).where(Equipment.is_active == True).order_by(Equipment.category, Equipment.name)
        )
        equipment = result.scalars().all()
        return [
            {
                "name": e.name,
                "category": e.category.value,
                "condition": e.condition.value,
                "qr_code": e.qr_code,
                "total_quantity": e.total_quantity,
                "available_quantity": e.available_quantity,
                "borrowed_quantity": e.total_quantity - e.available_quantity,
                "storage_location": e.storage_location,
                "sport_or_art": e.sport_or_art,
            }
            for e in equipment
        ]

    def _render_attendance_html(self, records: list[dict], sport_or_art, date_from, date_to) -> str:
        rows_html = ""
        for i, r in enumerate(records):
            bg = "#EBF0F7" if i % 2 == 0 else "#FFFFFF"
            time_in = r["time_in"].strftime("%Y-%m-%d %H:%M") if r["time_in"] else "-"
            time_out = r["time_out"].strftime("%H:%M") if r["time_out"] else "-"
            duration = f"{r['duration_minutes']} min" if r["duration_minutes"] else "-"
            rows_html += f"""
            <tr style="background:{bg}">
                <td>{r['student_name']}</td>
                <td>{r['student_id'] or '-'}</td>
                <td>{r['session_name']}</td>
                <td>{r['sport_or_art'] or '-'}</td>
                <td>{r['activity_type']}</td>
                <td>{time_in}</td>
                <td>{time_out}</td>
                <td>{duration}</td>
                <td>{'✓' if r['is_complete'] else '⏳'}</td>
            </tr>"""

        filter_info = f"Sport/Art: {sport_or_art or 'All'} | Period: {date_from or 'All'} – {date_to or 'All'}"
        return f"""
        <!DOCTYPE html><html><head>
        <meta charset="UTF-8">
        <style>
            body {{ font-family: Arial, sans-serif; font-size: 10pt; margin: 20px; }}
            h1 {{ color: #1E3A5F; }} h2 {{ color: #666; font-size: 10pt; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
            th {{ background: #1E3A5F; color: white; padding: 6px 8px; text-align: left; font-size: 9pt; }}
            td {{ padding: 5px 8px; border-bottom: 1px solid #DDD; font-size: 9pt; }}
            .footer {{ margin-top: 20px; font-size: 8pt; color: #999; }}
        </style></head><body>
        <h1>NAAP-Villamor OSCA — Attendance Report</h1>
        <h2>{filter_info}</h2>
        <p>Total records: {len(records)} | Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}</p>
        <table>
        <tr>
            <th>Student Name</th><th>Student ID</th><th>Session</th><th>Sport/Art</th>
            <th>Activity</th><th>Time In</th><th>Time Out</th><th>Duration</th><th>Complete</th>
        </tr>
        {rows_html}
        </table>
        <div class="footer">
            OSCA Attendance & Inventory Management System v2.0 — Confidential
        </div>
        </body></html>"""

    def _render_inventory_html(self, equipment: list[dict]) -> str:
        rows_html = ""
        for i, e in enumerate(equipment):
            bg = "#EBF0F7" if i % 2 == 0 else "#FFFFFF"
            rows_html += f"""
            <tr style="background:{bg}">
                <td>{e['name']}</td><td>{e['category']}</td><td>{e['condition']}</td>
                <td>{e['qr_code']}</td><td>{e['total_quantity']}</td>
                <td>{e['available_quantity']}</td><td>{e['borrowed_quantity']}</td>
                <td>{e['storage_location'] or '-'}</td><td>{e['sport_or_art'] or '-'}</td>
            </tr>"""

        return f"""
        <!DOCTYPE html><html><head><meta charset="UTF-8">
        <style>
            body {{ font-family: Arial, sans-serif; font-size: 10pt; margin: 20px; }}
            h1 {{ color: #1E3A5F; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
            th {{ background: #1E3A5F; color: white; padding: 6px 8px; text-align: left; }}
            td {{ padding: 5px 8px; border-bottom: 1px solid #DDD; font-size: 9pt; }}
        </style></head><body>
        <h1>NAAP-Villamor OSCA — Equipment Inventory Report</h1>
        <p>Total items: {len(equipment)} | Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}</p>
        <table>
        <tr>
            <th>Equipment Name</th><th>Category</th><th>Condition</th><th>QR Code</th>
            <th>Total Qty</th><th>Available</th><th>Borrowed</th><th>Location</th><th>Sport/Art</th>
        </tr>
        {rows_html}
        </table>
        </body></html>"""

    def _render_monthly_html(self, report: dict) -> str:
        period = report["period"]
        import calendar
        month_name = calendar.month_name[period["month"]]

        top_rows = ""
        for i, e in enumerate(report["top_5_borrowed"]):
            bg = "#EBF0F7" if i % 2 == 0 else "#FFFFFF"
            top_rows += f'<tr style="background:{bg}"><td>{e["name"]}</td><td>{e["borrow_count"]}</td></tr>'

        condition_rows = ""
        for cond, count in report["condition_breakdown"].items():
            condition_rows += f"<tr><td>{cond.title()}</td><td>{count}</td></tr>"

        return f"""
        <!DOCTYPE html><html><head><meta charset="UTF-8">
        <style>
            body {{ font-family: Arial, sans-serif; font-size: 11pt; margin: 20px; }}
            h1 {{ color: #1E3A5F; }} h2 {{ color: #1E3A5F; font-size: 12pt; margin-top: 20px; }}
            .stat-grid {{ display: flex; gap: 16px; flex-wrap: wrap; margin: 12px 0; }}
            .stat-box {{ background: #EBF0F7; border-radius: 8px; padding: 12px 20px; min-width: 140px; }}
            .stat-label {{ font-size: 9pt; color: #666; }} .stat-value {{ font-size: 18pt; font-weight: bold; color: #1E3A5F; }}
            table {{ border-collapse: collapse; margin-top: 8px; }} th {{ background: #1E3A5F; color: white; padding: 6px 14px; }}
            td {{ padding: 5px 14px; border-bottom: 1px solid #DDD; }}
        </style></head><body>
        <h1>Monthly Inventory Summary — {month_name} {period["year"]}</h1>
        <p>Generated: {report["generated_at"]}</p>
        <div class="stat-grid">
            <div class="stat-box"><div class="stat-label">Active Equipment</div>
                <div class="stat-value">{report["total_active_equipment"]}</div></div>
            <div class="stat-box"><div class="stat-label">Borrowed This Month</div>
                <div class="stat-value">{report["borrowed_this_month"]}</div></div>
            <div class="stat-box"><div class="stat-label">Returned This Month</div>
                <div class="stat-value">{report["returned_this_month"]}</div></div>
            <div class="stat-box"><div class="stat-label">Overdue</div>
                <div class="stat-value" style="color:#E53E3E">{report["overdue_at_end_of_month"]}</div></div>
        </div>
        <h2>Top 5 Most Borrowed</h2>
        <table><tr><th>Equipment</th><th>Borrow Count</th></tr>{top_rows}</table>
        <h2>Equipment Condition Breakdown</h2>
        <table><tr><th>Condition</th><th>Count</th></tr>{condition_rows}</table>
        </body></html>"""

    def _html_to_pdf(self, html: str) -> bytes:
        from xhtml2pdf import pisa
        buffer = io.BytesIO()
        pisa.CreatePDF(io.StringIO(html), dest=buffer)
        return buffer.getvalue()

    def _build_attendance_xlsx(self, records: list[dict]) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        headers = ["Student Name", "Student ID", "Session", "Sport/Art",
                   "Activity", "Time In", "Time Out", "Duration (min)", "Complete"]
        self._write_header_row(ws, headers)

        for i, r in enumerate(records, start=2):
            row_fill = ALT_FILL if i % 2 == 0 else None
            values = [
                r["student_name"], r["student_id"], r["session_name"],
                r["sport_or_art"], r["activity_type"],
                r["time_in"], r["time_out"], r["duration_minutes"],
                "Yes" if r["is_complete"] else "No",
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=i, column=col, value=val)
                if row_fill:
                    cell.fill = row_fill

        self._auto_column_width(ws)
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _build_inventory_xlsx(self, equipment: list[dict]) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory Report"

        headers = ["Equipment Name", "Category", "Condition", "QR Code",
                   "Total Qty", "Available", "Borrowed", "Location", "Sport/Art"]
        self._write_header_row(ws, headers)

        for i, e in enumerate(equipment, start=2):
            values = [
                e["name"], e["category"], e["condition"], e["qr_code"],
                e["total_quantity"], e["available_quantity"], e["borrowed_quantity"],
                e["storage_location"], e["sport_or_art"],
            ]
            row_fill = ALT_FILL if i % 2 == 0 else None
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=i, column=col, value=val)
                if row_fill:
                    cell.fill = row_fill

        self._auto_column_width(ws)
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _build_monthly_xlsx(self, report: dict) -> bytes:
        wb = openpyxl.Workbook()

        # Summary sheet
        ws = wb.active
        ws.title = "Summary"
        period = report["period"]
        import calendar
        ws["A1"] = f"Monthly Inventory Summary — {calendar.month_name[period['month']]} {period['year']}"
        ws["A1"].font = Font(bold=True, size=14, color="1E3A5F")

        summary_rows = [
            ("Active Equipment", report["total_active_equipment"]),
            ("Borrowed This Month", report["borrowed_this_month"]),
            ("Returned This Month", report["returned_this_month"]),
            ("Overdue at End of Month", report["overdue_at_end_of_month"]),
        ]
        for row_idx, (label, value) in enumerate(summary_rows, start=3):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value)

        # Top 5 sheet
        ws2 = wb.create_sheet("Top 5 Borrowed")
        self._write_header_row(ws2, ["Equipment Name", "Borrow Count"])
        for i, e in enumerate(report["top_5_borrowed"], start=2):
            ws2.cell(row=i, column=1, value=e["name"])
            ws2.cell(row=i, column=2, value=e["borrow_count"])

        # Condition sheet
        ws3 = wb.create_sheet("Condition Breakdown")
        self._write_header_row(ws3, ["Condition", "Count"])
        for i, (cond, count) in enumerate(report["condition_breakdown"].items(), start=2):
            ws3.cell(row=i, column=1, value=cond.title())
            ws3.cell(row=i, column=2, value=count)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _write_header_row(self, ws, headers: list[str]) -> None:
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20

    def _auto_column_width(self, ws) -> None:
        for column in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in column), default=10)
            ws.column_dimensions[column[0].column_letter].width = min(max_len + 4, 50)
